"""Epson Korea MI Platform — FastAPI backend.

설계 단계: Salesforce 실연동 없이 data/sample/*.json을 시드로 하는
인메모리 스토어를 서빙한다. 엑셀 업로드는 스토어에 반영되고(서버 재시작 시
시드로 초기화), Export는 현재 스토어를 Salesforce Data Loader 포맷으로 변환한다.

실행:  uvicorn backend.main:app --reload --port 8000  (저장소 루트에서)
"""
import io
import json
import re
from datetime import date, datetime
from pathlib import Path

from fastapi import FastAPI, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles

from backend import similarity

ROOT = Path(__file__).resolve().parent.parent
SAMPLE_DIR = ROOT / "data" / "sample"
DIST_DIR = ROOT / "frontend" / "dist"

app = FastAPI(title="Epson Korea MI Platform API", version="0.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_methods=["GET", "POST", "PATCH"],
    allow_headers=["*"],
)


def _load(name: str):
    path = SAMPLE_DIR / f"{name}.json"
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"sample data not found: {name}")
    with path.open(encoding="utf-8") as f:
        return json.load(f)


# ── 인메모리 스토어 (시드: data/sample) ─────────────────────────
STORE = {
    "opportunities": _load("opportunities"),
    "accounts": _load("accounts"),
    "sales_plan": _load("sales_plan"),
    "sensing_events": _load("sensing_events"),
    "internal_resources": _load("internal_resources"),
    "tech_signals": _load("tech_signals"),
    "tech_trends": _load("tech_trends"),
}

BUS = {"PRT", "PJT", "RBT", "CMP"}
CHANNELS = {"직판", "총판"}
STAGE_PROB = {s["stage"]: s["probability"] for s in STORE["opportunities"]["stageDefinition"]}


@app.get("/api/opportunities")
def opportunities():
    return STORE["opportunities"]


@app.get("/api/accounts")
def accounts():
    return STORE["accounts"]


@app.get("/api/sales-plan")
def sales_plan():
    return STORE["sales_plan"]


@app.get("/api/sensing-events")
def sensing_events():
    return STORE["sensing_events"]


@app.get("/api/health")
def health():
    return {"status": "ok"}


# ── 승격·상태 변경 저장 ────────────────────────────────────────
@app.post("/api/opportunities")
def create_opportunity(opp: dict):
    opps = STORE["opportunities"]["opportunities"]
    oid = opp.get("opportunityId")
    if not oid or not opp.get("name") or not opp.get("amount"):
        raise HTTPException(status_code=422, detail="opportunityId, name, amount는 필수입니다")
    existing = next((o for o in opps if o["opportunityId"] == oid), None)
    if existing:
        existing.update(opp)
    else:
        opps.append(opp)
    return {"opportunityId": oid, "updated": bool(existing)}


@app.patch("/api/sensing-events/{event_id}")
def update_event(event_id: str, patch: dict):
    allowed = {"status", "promotedOpportunityId", "assignedOwner"}
    ev = next((e for e in STORE["sensing_events"]["events"] if e["eventId"] == event_id), None)
    if not ev:
        raise HTTPException(status_code=404, detail=f"event not found: {event_id}")
    ev.update({k: v for k, v in patch.items() if k in allowed})
    return ev


# ── 대응 적합성 판단 (승격 전 내부 리소스·유사사례 근거 제시) ──
WON_STAGES = {"계약 완료(수주)", "매출 인식"}
LOST_STAGE = "실주(Lost)"
BU_NAMES = {"PRT": "프린팅", "PJT": "프로젝터", "RBT": "로봇", "CMP": "부품·소재"}


def _opp_outcome(o):
    if o["stage"] in WON_STAGES:
        return "수주"
    if o["stage"] == LOST_STAGE:
        return "실주"
    return "진행중"


def _opp_doc(o):
    return " ".join(
        str(o.get(k) or "")
        for k in ("name", "productFamily", "productModel", "customerSegment", "painPoint", "description")
    )


def _clamp(v):
    return max(0, min(100, round(v)))


def _active_load(owner_name):
    active = [
        o for o in STORE["opportunities"]["opportunities"]
        if o.get("owner") == owner_name and o["stage"] not in WON_STAGES and o["stage"] != LOST_STAGE
    ]
    return {"count": len(active), "amount": sum(o.get("amount") or 0 for o in active)}


def _family_match(fam_a, fam_b):
    """'산업용 잉크젯, 6축 로봇' 같은 복수 표기 대비 토큰 겹침으로 제품군을 느슨하게 매칭."""
    ta = set(re.findall(r"[가-힣A-Za-z0-9]+", str(fam_a or "")))
    tb = set(re.findall(r"[가-힣A-Za-z0-9]+", str(fam_b or "")))
    return bool(ta & tb)


@app.get("/api/fit-assessment/{event_id}")
def fit_assessment(event_id: str):
    """승격 전 판단 근거: 과거 유사사례(임베딩 유사도) + 내부 리소스 가용성 4개 축."""
    ev = next((e for e in STORE["sensing_events"]["events"] if e["eventId"] == event_id), None)
    if not ev:
        raise HTTPException(status_code=404, detail=f"event not found: {event_id}")

    res = STORE["internal_resources"]
    opps = [
        o for o in STORE["opportunities"]["opportunities"]
        if o.get("sensingEventId") != event_id  # 이 이벤트에서 승격된 건 자기 자신이므로 제외
    ]
    bu = (ev.get("estimatedBusinessUnit") or ["PRT"])[0]
    fam = ev.get("estimatedProductFamily") or ""
    axes = []
    cautions = []

    # 1) 유사사례 실적 — 임베딩(폴백 TF-IDF) 유사도 Top-3 × 결과 가중
    query = f"{ev['targetName']} {ev['summary']} {fam} {ev['triggerType']}"
    method, ranked = similarity.rank(query, [_opp_doc(o) for o in opps])
    similar_cases = []
    for idx, score in ranked[:3]:
        o = opps[idx]
        matched_on = []
        if o.get("businessUnit") == bu:
            matched_on.append(f"사업부 {BU_NAMES.get(bu, bu)}")
        if _family_match(o.get("productFamily"), fam):
            matched_on.append("제품군 유사")
        if ev["category"] == "B2G" and o.get("customerSegment") == "B2G공공":
            matched_on.append("공공(B2G) 세그먼트")
        similar_cases.append({
            "opportunityId": o["opportunityId"],
            "name": o["name"],
            "outcome": _opp_outcome(o),
            "stage": o["stage"],
            "amount": o.get("amount"),
            "owner": o.get("owner"),
            "similarity": round(score, 3),
            "matchedOn": matched_on,
        })
    best_sim = similar_cases[0]["similarity"] if similar_cases else 0.0
    won_cnt = sum(1 for c in similar_cases if c["outcome"] == "수주")
    lost_cnt = sum(1 for c in similar_cases if c["outcome"] == "실주")
    sim_score = _clamp(best_sim * 100 + 12 * min(won_cnt, 2) - 15 * lost_cnt)
    sim_reasons = []
    if similar_cases:
        top = similar_cases[0]
        sim_reasons.append(f"최고 유사 사례 {top['opportunityId']} ({top['outcome']}) — 유사도 {round(top['similarity'] * 100)}%")
        if won_cnt:
            sim_reasons.append(f"유사 Top3 중 수주 성공 {won_cnt}건 → 대응 경험 보유")
        if lost_cnt:
            sim_reasons.append(f"유사 Top3 중 실주 {lost_cnt}건 → 실패 요인 사전 점검 필요")
    else:
        sim_reasons.append("비교할 과거 사업기회가 없습니다")
    # 동일 사업부의 경쟁사 낙찰·스펙락인 이력은 주의사항으로
    for other in STORE["sensing_events"]["events"]:
        d = other.get("detail") or {}
        if (
            other["eventId"] != event_id
            and d.get("type") == "bid"
            and d.get("awardedVendor")
            and bu in (other.get("estimatedBusinessUnit") or [])
        ):
            cautions.append(
                f"동일 사업부 최근 경쟁사 낙찰 이력: {d['awardedVendor']} ({d['orderingOrg']}, 낙찰률 {d.get('awardRate')}%)"
                + (" — 스펙락인 정황 주의" if str(d.get("competitorSpecLockIn", "")).startswith("예") else "")
            )
    axes.append({"key": "similarCases", "label": "유사사례 실적", "score": sim_score, "reasons": sim_reasons})

    # 2) 영업 가용성 — 사업부 담당 후보의 진행중 로드(실시간 계산)
    candidates = [r for r in res["salesReps"] if bu in r["businessUnits"]] or res["salesReps"]
    loads = [(r, _active_load(r["name"])) for r in candidates]
    loads.sort(key=lambda x: (x[1]["count"], x[1]["amount"]))
    best_rep, best_load = loads[0]
    assigned = ev.get("assignedOwner")
    base_load = next((l for r, l in loads if r["name"] == assigned), best_load) if assigned else best_load
    sales_score = _clamp(95 - 14 * base_load["count"])
    sales_reasons = [
        f"{assigned or best_rep['name']} 진행중 {base_load['count']}건 ({base_load['amount'] / 1e8:.1f}억) 담당중"
    ]
    if not assigned:
        sales_reasons.append(f"미배정 — {BU_NAMES.get(bu, bu)} 담당 중 로드 최소는 {best_rep['name']}")
    if base_load["count"] >= 5:
        cautions.append("담당 영업 로드가 높습니다 — 배정 조정 또는 우선순위 검토 권장")
    axes.append({"key": "salesCapacity", "label": "영업 가용성", "score": sales_score, "reasons": sales_reasons})

    # 3) 기술지원·데모 여력 — SE 가동률 + 데모 장비 잔여 재고
    se = next((t for t in res["techSupport"] if t["businessUnit"] == bu), None)
    demo = next((d for d in res["demoEquipment"] if _family_match(d["productFamily"], fam)), None)
    tech_score = 100 - (se["utilization"] if se else 50)
    tech_reasons = []
    if se:
        tech_reasons.append(f"{BU_NAMES.get(bu, bu)} SE {se['seHeadcount']}명 · 가동률 {se['utilization']}% · 동시 PoC {se['maxConcurrentPoc']}건 가능")
        if se["utilization"] >= 80:
            cautions.append(f"{BU_NAMES.get(bu, bu)} 기술지원 가동률 {se['utilization']}% — Demo·PoC 일정 지연 가능")
    if demo:
        remain = demo["total"] - demo["onLoan"]
        tech_reasons.append(f"데모 장비 {demo['productFamily']}({demo['model']}) 잔여 {remain}대 / 보유 {demo['total']}대")
        if remain <= 0:
            tech_score -= 20
            cautions.append(f"데모 장비({demo['productFamily']}) 전량 대여중 — 회수 일정 확인 필요")
        else:
            tech_score += 10
    d = ev.get("detail") or {}
    if d.get("type") == "bid" and not d.get("awardDate"):
        try:
            days = (datetime.strptime(d["bidDeadline"], "%Y-%m-%d").date() - date(2026, 7, 29)).days
            if 0 <= days <= 14:
                tech_score -= 10
                cautions.append(f"입찰 마감 D-{days} — 규격서 검토·제안 준비 기간 촉박")
        except (KeyError, ValueError):
            pass
    axes.append({"key": "techCapacity", "label": "기술지원·데모 여력", "score": _clamp(tech_score), "reasons": tech_reasons})

    # 4) 채널·파트너 커버리지
    account = next((a for a in STORE["accounts"]["accounts"] if a["accountId"] == ev.get("matchedAccountId")), None)
    partner = next(
        (p for p in res["partners"] if bu in p["businessUnits"] and (not fam or any(_family_match(pf, fam) for pf in p["productFamilies"]))),
        None,
    )
    ch_reasons = []
    recommended_partner = None
    if account:
        ch_score = 85
        ch_reasons.append(f"기존 계정 {account['accountName']} — 채널 {account['channel']} 거래 이력 보유")
        if account["channel"] == "총판" and partner:
            recommended_partner = {"name": partner["name"], "reason": f"{'/'.join(partner['productFamilies'][:2])} 취급 · 커버 지역 {'/'.join(partner['regions'])}"}
            ch_score = 90
    elif partner:
        ch_score = 65
        ch_reasons.append(f"신규 기관 — 총판 {partner['name']} 경유 대응 가능 ({'/'.join(partner['regions'])})")
        recommended_partner = {"name": partner["name"], "reason": f"{BU_NAMES.get(bu, bu)} 제품군 취급 파트너"}
    else:
        ch_score = 50
        ch_reasons.append("신규 기관 — 해당 제품군 커버 파트너 없음, 직판 신규 개척 필요")
    axes.append({"key": "channelFit", "label": "채널·파트너 커버리지", "score": _clamp(ch_score), "reasons": ch_reasons})

    weights = {"similarCases": 0.35, "salesCapacity": 0.25, "techCapacity": 0.25, "channelFit": 0.15}
    fit_score = _clamp(sum(a["score"] * weights[a["key"]] for a in axes))
    recommendation = "승격 권장" if fit_score >= 70 else "조건부 승격" if fit_score >= 40 else "보류 권장"

    return {
        "eventId": event_id,
        "recommendation": recommendation,
        "fitScore": fit_score,
        "method": method,
        "methodLabel": "의미 임베딩 (multilingual-e5-small)" if method == "embedding" else "키워드 유사도 (TF-IDF 폴백)",
        "axes": axes,
        "similarCases": similar_cases,
        "recommendedOwner": {
            "name": best_rep["name"],
            "reason": f"{best_rep['specialty']} · 진행중 {best_load['count']}건으로 로드 최소",
            "currentLoad": best_load,
        },
        "recommendedPartner": recommended_partner,
        "cautions": cautions,
    }


# ── 기술 트렌드 팔로업 (원시 시그널 → 트렌드 클러스터) ────────
@app.get("/api/tech-signals")
def tech_signals():
    return STORE["tech_signals"]


@app.get("/api/tech-trends")
def tech_trends():
    return STORE["tech_trends"]


# 상태 전이: 신규 →[추적 시작]→ 추적중 →[보고 처리]→ 보고완료 / 어디서든 [중단] → 중단 →[재개]→ 추적중
TREND_STATUSES = {"신규", "추적중", "보고완료", "중단"}


@app.patch("/api/tech-trends/{trend_id}")
def update_trend(trend_id: str, patch: dict):
    trend = next((t for t in STORE["tech_trends"]["trends"] if t["trendId"] == trend_id), None)
    if not trend:
        raise HTTPException(status_code=404, detail=f"trend not found: {trend_id}")
    status = patch.get("status")
    if status not in TREND_STATUSES:
        raise HTTPException(status_code=422, detail=f"status는 {'/'.join(sorted(TREND_STATUSES))} 중 하나여야 합니다")
    trend["status"] = status
    return trend


@app.get("/api/tech-trends/{trend_id}/related")
def tech_trend_related(trend_id: str):
    """트렌드 ↔ 우리 파이프라인 연결: 사업부가 겹치는 사업기회·센싱 이벤트를
    유사도(임베딩, 폴백 TF-IDF)로 매칭한다. 실서비스에서는 담당자 수동 확정 링크 병행."""
    trend = next((t for t in STORE["tech_trends"]["trends"] if t["trendId"] == trend_id), None)
    if not trend:
        raise HTTPException(status_code=404, detail=f"trend not found: {trend_id}")

    bus = set(trend["businessUnit"])
    query = f"{trend['title']} {trend['summary']} {trend['techCategory']} {' '.join(trend['tags'])}"

    opps = [
        o for o in STORE["opportunities"]["opportunities"]
        if o.get("businessUnit") in bus and o["stage"] != LOST_STAGE
    ]
    method, ranked = similarity.rank(query, [_opp_doc(o) for o in opps])
    rel_opps = [
        {
            "opportunityId": opps[i]["opportunityId"],
            "name": opps[i]["name"],
            "stage": opps[i]["stage"],
            "outcome": _opp_outcome(opps[i]),
            "amount": opps[i].get("amount"),
            "owner": opps[i].get("owner"),
            "similarity": round(s, 3),
        }
        for i, s in ranked[:3]
        if s >= similarity.RELATED_FLOOR[method]
    ]

    evts = [
        e for e in STORE["sensing_events"]["events"]
        if bus & set(e.get("estimatedBusinessUnit") or [])
    ]
    method2, ranked2 = similarity.rank(
        query, [f"{e['targetName']} {e['summary']} {e.get('estimatedProductFamily') or ''}" for e in evts]
    )
    rel_evts = [
        {
            "eventId": evts[i]["eventId"],
            "targetName": evts[i]["targetName"],
            "summary": evts[i]["summary"],
            "status": evts[i]["status"],
            "promotedOpportunityId": evts[i].get("promotedOpportunityId"),
            "similarity": round(s, 3),
        }
        for i, s in ranked2[:3]
        if s >= similarity.RELATED_FLOOR[method2]
    ]

    return {
        "trendId": trend_id,
        "method": method,
        "opportunities": rel_opps,
        "events": rel_evts,
        "pipelineAmount": sum(o["amount"] or 0 for o in rel_opps),
    }


# ── 엑셀 업로드 (실무 양식 → 스토어 반영) ──────────────────────
# 컬럼명은 excel/sales_pipeline_template.xlsx '사업기회' 시트 헤더와 1:1
OPP_COLS = {
    "사업기회ID": "opportunityId",
    "사업기회명": "name",
    "고객사명": "accountName",
    "사업부": "businessUnit",
    "채널": "channel",
    "파트너사": "partnerAccount",
    "고객유형": "customerSegment",
    "제품군": "productFamily",
    "대표 모델": "productModel",
    "예상 수량": "quantity",
    "예상 금액(원)": "amount",
    "단계": "stage",
    "확도(%)": "probability",
    "예상 수주일": "closeDate",
    "고객 Pain Point": "painPoint",
    "요구 조건(Target Spec)": "targetSpec",
    "Demo·PoC 목표": "demoTargetDate",
    "검증(QUAL) 목표": "qualTargetDate",
    "경쟁사": "competitor",
    "관련 사업기회": "relatedOpportunity",
    "유입 경로": "leadSource",
    "담당 영업": "owner",
    "비고": "description",
}
PLAN_COLS = {"연도": "year", "월": "month", "사업부": "businessUnit", "채널": "channel", "목표 금액(원)": "targetAmount"}
REQUIRED = ["name", "businessUnit", "channel", "stage", "amount", "closeDate"]
DATE_FIELDS = ["closeDate", "demoTargetDate", "qualTargetDate"]


def _to_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    raise ValueError(f"날짜 형식 오류(YYYY-MM-DD): {v}")


def _next_opp_id(opps):
    mx = 0
    for o in opps:
        m = re.search(r"(\d+)$", o["opportunityId"])
        if m:
            mx = max(mx, int(m.group(1)))
    return f"OPP-2026-{mx + 1:04d}"


def _header_map(ws, col_def):
    headers = [c.value for c in ws[1]]
    return {idx: col_def[h] for idx, h in enumerate(headers) if h in col_def}



# ── 현업 자유양식 엑셀 매핑 (컬럼 유사어·단위·단계 용어·날짜 표기 해석) ──
FIELD_HEADER_SYNONYMS = {
    "name": ["건명", "사업명", "안건", "내용", "프로젝트", "사업기회명"],
    "accountName": ["업체명", "업체", "거래처", "거래처명", "고객사", "고객사명", "고객명", "기관명"],
    "productModel": ["품목/모델", "품목", "모델", "제품", "제품명", "장비"],
    "quantity": ["수량", "대수", "예상 수량"],
    "amount": ["예상금액", "금액", "예상 금액", "수주예상액", "규모", "예상금액(백만원)", "금액(백만)", "예상 금액(원)"],
    "stage": ["진행상태", "상태", "진행", "단계", "진행단계", "진척"],
    "closeDate": ["수주예정", "수주예정일", "납기", "예정일", "수주시기", "예상 수주일"],
    "owner": ["담당", "담당자", "영업담당", "담당 영업"],
    "description": ["비고", "메모", "특이사항", "코멘트"],
}

STAGE_SYNONYMS = [
    (["리드", "발굴", "신규", "타겟"], "리드 발굴"),
    (["상담", "니즈", "미팅", "접촉"], "니즈 파악·상담"),
    (["견적", "제안", "입찰준비"], "제안·견적"),
    (["데모", "demo", "poc", "테스트", "실증", "시연"], "Demo·PoC/테스트"),
    (["협상", "계약대기", "계약검토", "품의"], "협상·계약검토"),
    (["수주", "계약완료", "계약 완료", "낙찰"], "계약 완료(수주)"),
    (["납품완료", "검수", "매출", "완료"], "매출 인식"),
    (["실주", "취소", "드랍", "lost"], "실주(Lost)"),
]

BU_KEYWORDS = [
    # 명시적 프린터 키워드를 먼저 확인 (모델명 부분 문자열 오탐 방지: WF-C879R의 'c8' 등)
    (["복합기", "프린터", "잉크젯", "라벨", "영수증", "mps", "wf-", "tm-", "am-", "cw-"], "PRT"),
    (["프로젝터", "eb-", "빔"], "PJT"),
    (["로봇", "scara", "6축", "gx8", "gx4", "c8xl", "c12xl"], "RBT"),
    (["분말", "헤드", "코어", "precisioncore", "atmix"], "CMP"),
]


def _map_stage(raw):
    s = str(raw).strip().lower()
    for keys, stage in STAGE_SYNONYMS:
        if any(k in s for k in keys):
            return stage
    return None


def _guess_bu(text):
    s = str(text).lower()
    for keys, bu in BU_KEYWORDS:
        if any(k in s for k in keys):
            return bu
    return "PRT"


def _parse_field_amount(v, header=""):
    """'0.9억' / '45,000,000' / 270(백만원 표기·소액 숫자) 등 현업 금액 표기를 원 단위로."""
    if v in (None, ""):
        raise ValueError("금액 누락")
    if isinstance(v, (int, float)):
        n = float(v)
        if "억" in header:
            return int(n * 1e8), None
        if "백만" in header or n < 100000:
            return int(n * 1e6), "금액을 백만원 단위로 해석"
        return int(n), None
    s = str(v).replace(",", "").replace(" ", "").replace("원", "")
    if s.endswith("억"):
        return int(float(s[:-1]) * 1e8), f"'{v}' → 억 단위 해석"
    if s.endswith("백만"):
        return int(float(s[:-2]) * 1e6), f"'{v}' → 백만원 단위 해석"
    if s.endswith("천만"):
        return int(float(s[:-2]) * 1e7), f"'{v}' → 천만원 단위 해석"
    n = float(s)
    if n < 100000:
        return int(n * 1e6), "금액을 백만원 단위로 해석"
    return int(n), None


def _parse_field_date(v):
    """'2026.10.31' / '12/15' / '11월말' / date 셀 등 → YYYY-MM-DD."""
    import calendar
    if v in (None, ""):
        return None, None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d"), None
    s = str(v).strip().replace(".", "-").replace("/", "-")
    m = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, mo, d = map(int, m.groups())
        return f"{y:04d}-{mo:02d}-{d:02d}", None
    m = re.fullmatch(r"(\d{1,2})-(\d{1,2})", s)
    if m:
        mo, d = map(int, m.groups())
        return f"2026-{mo:02d}-{d:02d}", f"'{v}' → 2026년으로 해석"
    m = re.fullmatch(r"(\d{1,2})월\s*(말|초|중순)?", str(v).strip())
    if m:
        mo = int(m.group(1))
        part = m.group(2) or "말"
        d = {"초": 5, "중순": 15}.get(part, calendar.monthrange(2026, mo)[1])
        return f"2026-{mo:02d}-{d:02d}", f"'{v}' → 2026-{mo:02d}-{d:02d}로 해석"
    raise ValueError(f"날짜 해석 불가: {v}")


def _find_field_sheet(wb):
    """유사 컬럼명이 3개 이상 잡히는 헤더 행을 가진 시트를 찾는다 (제목행 허용, 1~5행 탐색)."""
    for ws in wb.worksheets:
        for hr in range(1, min(6, ws.max_row + 1)):
            headers = [str(c.value).strip() if c.value is not None else "" for c in ws[hr]]
            colmap = {}
            for idx, h in enumerate(headers):
                for key, names in FIELD_HEADER_SYNONYMS.items():
                    if h in names and key not in colmap:
                        colmap[idx] = key
            if len(colmap) >= 3 and "name" in colmap.values():
                return ws, hr, colmap
    return None, None, None


STAGING = {}         # batchId → 미리보기(미반영) 배치
UPLOAD_HISTORY = []  # 반영된 배치 (롤백용 undo 스냅샷 포함)


def _parse_workbook(wb, filename):
    """엑셀을 파싱·검증만 하고 변경 예정 내역(미리보기)을 만든다. 스토어는 건드리지 않는다."""
    import uuid

    preview = {
        "batchId": uuid.uuid4().hex[:12],
        "filename": filename,
        "opportunities": {"inserts": [], "updates": []},
        "plan": {"changes": []},
        "errors": [],
    }
    opps = STORE["opportunities"]["opportunities"]
    acc_by_name = {a["accountName"]: a["accountId"] for a in STORE["accounts"]["accounts"]}
    staged_new = 0

    if "사업기회" in wb.sheetnames:
        ws = wb["사업기회"]
        hmap = _header_map(ws, OPP_COLS)
        if "name" not in hmap.values():
            preview["errors"].append({"sheet": "사업기회", "row": 1, "message": "헤더가 양식과 다릅니다 (입력가이드 시트 참조)"})
        else:
            for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
                vals = {hmap[i]: c.value for i, c in enumerate(row) if i in hmap}
                if not any(v not in (None, "") for v in vals.values()):
                    continue  # 빈 행
                try:
                    missing = [k for k in REQUIRED if vals.get(k) in (None, "")]
                    if missing:
                        raise ValueError(f"필수값 누락: {', '.join(missing)}")
                    if vals["businessUnit"] not in BUS:
                        raise ValueError(f"사업부 코드 오류: {vals['businessUnit']} (PRT/PJT/RBT/CMP)")
                    if vals["channel"] not in CHANNELS:
                        raise ValueError(f"채널 오류: {vals['channel']} (직판/총판)")
                    if vals["stage"] not in STAGE_PROB and vals["stage"] != "실주(Lost)":
                        raise ValueError(f"단계 오류: {vals['stage']}")
                    amount = float(str(vals["amount"]).replace(",", ""))
                    for f_ in DATE_FIELDS:
                        vals[f_] = _to_date(vals.get(f_))
                    name = str(vals.get("accountName") or "").strip()
                    rec = {
                        **{v: None for v in OPP_COLS.values()},
                        **{k: (str(v).strip() if isinstance(v, str) else v) for k, v in vals.items()},
                        "amount": int(amount),
                        "quantity": int(vals["quantity"]) if vals.get("quantity") not in (None, "") else None,
                        "probability": int(vals["probability"]) if vals.get("probability") not in (None, "") else STAGE_PROB.get(vals["stage"], 0),
                        "accountId": acc_by_name.get(name),
                        "sensingEventId": None,
                    }
                    rec.pop("accountName", None)
                    oid = str(vals.get("opportunityId") or "").strip()
                    existing = next((o for o in opps if o["opportunityId"] == oid), None) if oid else None
                    if existing:
                        after = {**existing, **{k: v for k, v in rec.items() if k != "opportunityId"}}
                        preview["opportunities"]["updates"].append({"before": dict(existing), "after": after})
                    else:
                        if not oid:
                            # 미리보기 단계에서 확정 채번: 기존 최대 번호 + 이 배치 내 신규 순번
                            staged_new += 1
                            base = int(_next_opp_id(opps).rsplit("-", 1)[1])
                            oid = f"OPP-2026-{base + staged_new - 1:04d}"
                        rec["opportunityId"] = oid
                        preview["opportunities"]["inserts"].append(rec)
                except (ValueError, TypeError) as e:
                    preview["errors"].append({"sheet": "사업기회", "row": r, "message": str(e)})

    if "판매계획" in wb.sheetnames:
        ws = wb["판매계획"]
        hmap = _header_map(ws, PLAN_COLS)
        plan_rows = STORE["sales_plan"]["monthlyPlan"]
        for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
            vals = {hmap[i]: c.value for i, c in enumerate(row) if i in hmap}
            if not any(v not in (None, "") for v in vals.values()):
                continue
            try:
                y, m = int(vals["year"]), int(vals["month"])
                bu, ch = vals["businessUnit"], vals["channel"]
                amt = int(float(str(vals["targetAmount"]).replace(",", "")))
                if bu not in BUS or ch not in CHANNELS or not 1 <= m <= 12:
                    raise ValueError(f"코드값 오류: {bu}/{ch}/{m}월")
                target = next(
                    (p for p in plan_rows if p["year"] == y and p["month"] == m and p["businessUnit"] == bu and p["channel"] == ch),
                    None,
                )
                before = target["targetAmount"] if target else None
                if before == amt:
                    continue  # 변경 없음
                preview["plan"]["changes"].append(
                    {"year": y, "month": m, "businessUnit": bu, "channel": ch, "before": before, "after": amt}
                )
            except (ValueError, TypeError, KeyError) as e:
                preview["errors"].append({"sheet": "판매계획", "row": r, "message": str(e)})

    return preview


def _strip_paren(s):
    return re.sub(r"\(.*?\)", "", s).strip()


def _parse_field_rows(grid, colmap, header_row, batch_id, filename):
    """확정된 컬럼 매핑으로 자유양식 행들을 표준 레코드로 해석한다."""
    preview = {
        "batchId": batch_id,
        "filename": filename,
        "opportunities": {"inserts": [], "updates": []},
        "plan": {"changes": []},
        "errors": [],
        "fieldMode": True,
    }
    accounts_list = STORE["accounts"]["accounts"]

    def _match_account(raw):
        n = raw.strip()
        for a in accounts_list:
            if a["accountName"] == n:
                return a
        for a in accounts_list:
            if _strip_paren(a["accountName"]) == _strip_paren(n):
                return a
        for a in accounts_list:
            if n and (n in a["accountName"] or _strip_paren(a["accountName"]) in n):
                return a
        return None

    header_txt = " ".join(header_row)
    staged_new = 0
    for r, row in enumerate(grid, start=2):
        vals = {key: (row[i] if i < len(row) else None) for i, key in colmap.items()}
        if not any(v not in (None, "") for v in vals.values()):
            continue
        try:
            name = str(vals.get("name") or "").strip()
            acct = str(vals.get("accountName") or "").strip()
            if not name or not acct:
                raise ValueError("건명·업체명은 필수입니다")
            notes = []
            amount, note = _parse_field_amount(vals.get("amount"), header_txt)
            if note:
                notes.append(note)
            stage_raw = vals.get("stage")
            stage = _map_stage(stage_raw) if stage_raw not in (None, "") else None
            if stage is None:
                raise ValueError(f"진행상태 해석 불가: {stage_raw} (예: 견적, 상담중, 데모 예정, 수주)")
            if str(stage_raw).strip() != stage:
                notes.append(f"단계 '{stage_raw}' → {stage}")
            close, note = _parse_field_date(vals.get("closeDate"))
            if note:
                notes.append(note)
            if close is None:
                close = "2026-12-31"
                notes.append("수주예정 미기재 → 2026-12-31로 가정")
            model = str(vals.get("productModel") or "").strip()
            account = _match_account(acct)
            bu = _guess_bu(f"{model} {name}")
            notes.append(f"사업부 {bu} 추정 (품목 기반)")
            desc_txt = str(vals.get("description") or "")
            partner = None
            if "총판" in desc_txt or "총판" in name:
                channel = "총판"
                pm = re.search(r"총판\s*\(([^)]+)\)", desc_txt)
                if pm:
                    partner = pm.group(1).strip()
                notes.append(f"비고의 '총판' 언급 → 채널 총판{f' (파트너 {partner})' if partner else ''}")
            elif account:
                channel = account["channel"]
                notes.append(f"기존 계정 매칭({account['accountId']}) → 채널 {channel}")
            else:
                channel = "직판"
                notes.append("신규 업체 → 채널 직판 가정")
            qty = vals.get("quantity")
            rec = {
                **{v: None for v in OPP_COLS.values()},
                "name": f"{acct} {name}" if acct not in name else name,
                "businessUnit": bu,
                "channel": channel,
                "partnerAccount": partner,
                "customerSegment": account.get("customerSegment") if account else "B2B기업",
                "productModel": model or None,
                "productFamily": None,
                "quantity": int(qty) if qty not in (None, "") else None,
                "amount": amount,
                "stage": stage,
                "probability": STAGE_PROB.get(stage, 0),
                "closeDate": close,
                "owner": str(vals.get("owner") or "").strip() or None,
                "description": desc_txt.strip() or None,
                "leadSource": "엑셀 업로드",
                "accountId": account["accountId"] if account else None,
                "sensingEventId": None,
            }
            rec.pop("accountName", None)
            existing = next(
                (o for o in STORE["opportunities"]["opportunities"] if o["name"] == rec["name"]), None
            )
            if existing:
                after = {**existing, **{k: v for k, v in rec.items() if v is not None and k != "opportunityId"}}
                preview["opportunities"]["updates"].append({"before": dict(existing), "after": after, "mapNotes": notes})
            else:
                staged_new += 1
                base = int(_next_opp_id(STORE["opportunities"]["opportunities"]).rsplit("-", 1)[1])
                rec["opportunityId"] = f"OPP-2026-{base + staged_new - 1:04d}"
                rec["mapNotes"] = notes
                preview["opportunities"]["inserts"].append(rec)
        except (ValueError, TypeError) as e:
            preview["errors"].append({"sheet": "업로드 시트", "row": r, "message": str(e)})
    return preview


# 자유양식 컬럼 매핑 대상 필드 (셀포 필드 대응 관계 포함)
TARGET_FIELDS = [
    {"key": "name", "label": "사업기회명", "sf": "Name", "required": True},
    {"key": "accountName", "label": "고객사명", "sf": "AccountId(이름→ID 매칭)", "required": True},
    {"key": "productModel", "label": "대표 모델/품목", "sf": "ProductModel__c", "required": False},
    {"key": "quantity", "label": "예상 수량", "sf": "Quantity__c", "required": False},
    {"key": "amount", "label": "예상 금액", "sf": "Amount", "required": True},
    {"key": "stage", "label": "단계(진행상태)", "sf": "StageName", "required": True},
    {"key": "closeDate", "label": "예상 수주일", "sf": "CloseDate", "required": False},
    {"key": "owner", "label": "담당 영업", "sf": "OwnerId(이름→ID 매칭)", "required": False},
    {"key": "description", "label": "비고", "sf": "Description", "required": False},
]


def _column_proposal(ws, hr, auto_colmap):
    """자유양식 시트의 컬럼 목록 + 자동 매핑 제안 + 샘플 값 3개."""
    cols = []
    for idx, cell in enumerate(ws[hr]):
        header = str(cell.value).strip() if cell.value is not None else ""
        if not header:
            continue
        samples = []
        for row in ws.iter_rows(min_row=hr + 1, max_row=min(hr + 8, ws.max_row)):
            v = row[idx].value if idx < len(row) else None
            if v not in (None, ""):
                s = v.strftime("%Y-%m-%d") if isinstance(v, (datetime, date)) else str(v)
                samples.append(s)
            if len(samples) >= 3:
                break
        cols.append({"index": idx, "header": header, "samples": samples, "suggested": auto_colmap.get(idx)})
    return cols


@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile):
    """1단계: 양식 감지. 표준 양식은 바로 행 미리보기, 자유양식은 컬럼 매핑 제안을 반환한다."""
    import uuid
    from openpyxl import load_workbook

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="xlsx 파일만 업로드할 수 있습니다")
    try:
        wb = load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="엑셀 파일을 열 수 없습니다 (양식: sales_pipeline_template.xlsx)")

    if "사업기회" in wb.sheetnames or "판매계획" in wb.sheetnames:
        preview = _parse_workbook(wb, file.filename)
        preview["step"] = "rows"
        STAGING[preview["batchId"]] = preview
        return preview

    ws, hr, auto_colmap = _find_field_sheet(wb)
    if ws is None:
        raise HTTPException(
            status_code=422,
            detail="해석 가능한 시트를 찾지 못했습니다. 표준 양식(사업기회/판매계획 시트) 또는 업체명·건명·금액·진행상태 컬럼이 있는 영업관리 엑셀을 올려주세요.",
        )
    batch_id = uuid.uuid4().hex[:12]
    # 셀 값 그리드를 보관해 매핑 확정 시 재파싱한다
    grid = [[c.value for c in row] for row in ws.iter_rows(min_row=hr + 1)]
    STAGING[batch_id] = {
        "batchId": batch_id,
        "filename": file.filename,
        "sheet": ws.title,
        "headerRow": [str(c.value).strip() if c.value is not None else "" for c in ws[hr]],
        "grid": grid,
        "step": "mapping",
    }
    return {
        "batchId": batch_id,
        "step": "mapping",
        "filename": file.filename,
        "sheet": ws.title,
        "columns": _column_proposal(ws, hr, auto_colmap),
        "targetFields": TARGET_FIELDS,
    }


@app.post("/api/upload-map/{batch_id}")
def upload_map(batch_id: str, body: dict):
    """2단계: 사용자가 확정한 컬럼 매핑(colmap: {열index: 필드key})으로 행을 파싱해 미리보기를 만든다."""
    staged = STAGING.get(batch_id)
    if not staged or staged.get("step") != "mapping":
        raise HTTPException(status_code=404, detail="매핑 대기 중인 배치가 없습니다 (다시 업로드해 주세요)")
    raw_map = body.get("colmap") or {}
    colmap = {}
    valid_keys = {f["key"] for f in TARGET_FIELDS}
    for k, v in raw_map.items():
        if v in valid_keys:
            colmap[int(k)] = v
    mapped = set(colmap.values())
    missing = [f["label"] for f in TARGET_FIELDS if f["required"] and f["key"] not in mapped]
    if missing:
        raise HTTPException(status_code=422, detail=f"필수 필드가 매핑되지 않았습니다: {', '.join(missing)}")

    preview = _parse_field_rows(
        staged["grid"], colmap, staged["headerRow"], batch_id, staged["filename"]
    )
    preview["step"] = "rows"
    preview["sheet"] = staged["sheet"]
    STAGING[batch_id] = preview
    return preview


EDITABLE_FIELDS = {
    "name", "businessUnit", "channel", "partnerAccount", "stage", "probability",
    "amount", "closeDate", "owner", "quantity", "productModel", "description",
}


def _apply_edits(batch, edits):
    """미리보기 화면에서 사용자가 직접 수정한 값을 스테이징 배치에 덮어쓴다."""
    def merge(target, patch):
        for k, v in patch.items():
            if k not in EDITABLE_FIELDS:
                continue
            if k == "businessUnit" and v not in BUS:
                raise HTTPException(status_code=422, detail=f"사업부 코드 오류: {v}")
            if k == "channel" and v not in CHANNELS:
                raise HTTPException(status_code=422, detail=f"채널 오류: {v}")
            if k == "stage" and v not in STAGE_PROB and v != "실주(Lost)":
                raise HTTPException(status_code=422, detail=f"단계 오류: {v}")
            target[k] = v
        if "stage" in patch and "probability" not in patch:
            target["probability"] = STAGE_PROB.get(patch["stage"], 0)

    for e in edits.get("inserts", []):
        rec = next((r for r in batch["opportunities"]["inserts"] if r["opportunityId"] == e.get("opportunityId")), None)
        if rec:
            merge(rec, e)
    for e in edits.get("updates", []):
        u = next((u for u in batch["opportunities"]["updates"] if u["after"]["opportunityId"] == e.get("opportunityId")), None)
        if u:
            merge(u["after"], e)
    dropped = set(edits.get("dropIds", []))
    if dropped:
        batch["opportunities"]["inserts"] = [r for r in batch["opportunities"]["inserts"] if r["opportunityId"] not in dropped]
        batch["opportunities"]["updates"] = [u for u in batch["opportunities"]["updates"] if u["after"]["opportunityId"] not in dropped]


@app.post("/api/upload-apply/{batch_id}")
def upload_apply(batch_id: str, edits: dict | None = None):
    """2단계: 스테이징된 배치를 (사용자 수정분 반영 후) 스토어에 적용하고 히스토리에 기록한다."""
    batch = STAGING.pop(batch_id, None)
    if not batch:
        raise HTTPException(status_code=404, detail="배치를 찾을 수 없습니다 (이미 반영됐거나 만료)")
    if edits:
        _apply_edits(batch, edits)
    opps = STORE["opportunities"]["opportunities"]
    plan_rows = STORE["sales_plan"]["monthlyPlan"]

    for rec in batch["opportunities"]["inserts"]:
        rec = {k: v for k, v in rec.items() if k != "mapNotes"}
        if not any(o["opportunityId"] == rec["opportunityId"] for o in opps):
            opps.append(rec)
    for u in batch["opportunities"]["updates"]:
        target = next((o for o in opps if o["opportunityId"] == u["after"]["opportunityId"]), None)
        if target:
            target.clear()
            target.update(u["after"])
    for c in batch["plan"]["changes"]:
        target = next(
            (p for p in plan_rows if p["year"] == c["year"] and p["month"] == c["month"]
             and p["businessUnit"] == c["businessUnit"] and p["channel"] == c["channel"]),
            None,
        )
        if target:
            target["targetAmount"] = c["after"]
        else:
            plan_rows.append({"year": c["year"], "month": c["month"], "businessUnit": c["businessUnit"],
                              "channel": c["channel"], "targetAmount": c["after"]})

    entry = {
        "batchId": batch["batchId"],
        "filename": batch["filename"],
        "appliedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "counts": {
            "inserted": len(batch["opportunities"]["inserts"]),
            "updated": len(batch["opportunities"]["updates"]),
            "plan": len(batch["plan"]["changes"]),
        },
        "undo": {
            "insertedIds": [r["opportunityId"] for r in batch["opportunities"]["inserts"]],
            "updateSnapshots": [u["before"] for u in batch["opportunities"]["updates"]],
            "planChanges": batch["plan"]["changes"],
        },
    }
    UPLOAD_HISTORY.append(entry)
    return {"batchId": entry["batchId"], "counts": entry["counts"], "appliedAt": entry["appliedAt"]}


@app.get("/api/upload-history")
def upload_history():
    return [{k: e[k] for k in ("batchId", "filename", "appliedAt", "counts")} for e in reversed(UPLOAD_HISTORY)]


@app.delete("/api/upload-history/{batch_id}")
def upload_rollback(batch_id: str):
    """반영 제거(롤백): 신규 건 삭제, 갱신 건 이전 상태 복원, 계획 이전 금액 복원."""
    entry = next((e for e in UPLOAD_HISTORY if e["batchId"] == batch_id), None)
    if not entry:
        raise HTTPException(status_code=404, detail="히스토리에 없는 배치입니다")
    opps = STORE["opportunities"]["opportunities"]
    plan_rows = STORE["sales_plan"]["monthlyPlan"]
    undo = entry["undo"]

    STORE["opportunities"]["opportunities"] = [
        o for o in opps if o["opportunityId"] not in set(undo["insertedIds"])
    ]
    opps = STORE["opportunities"]["opportunities"]
    for snap in undo["updateSnapshots"]:
        target = next((o for o in opps if o["opportunityId"] == snap["opportunityId"]), None)
        if target:
            target.clear()
            target.update(snap)
    for c in undo["planChanges"]:
        target = next(
            (p for p in plan_rows if p["year"] == c["year"] and p["month"] == c["month"]
             and p["businessUnit"] == c["businessUnit"] and p["channel"] == c["channel"]),
            None,
        )
        if target:
            if c["before"] is None:
                plan_rows.remove(target)
            else:
                target["targetAmount"] = c["before"]

    UPLOAD_HISTORY.remove(entry)
    return {"removed": batch_id, "counts": entry["counts"]}


# ── Salesforce Data Loader 포맷 Export ────────────────────────
SF_COLS = [
    ("Name", "name"), ("AccountId", "accountId"), ("BusinessUnit__c", "businessUnit"),
    ("Channel__c", "channel"), ("PartnerAccount__c", "partnerAccount"),
    ("CustomerSegment__c", "customerSegment"), ("ProductFamily__c", "productFamily"),
    ("ProductModel__c", "productModel"), ("Quantity__c", "quantity"), ("Amount", "amount"),
    ("StageName", "stage"), ("Probability", "probability"), ("CloseDate", "closeDate"),
    ("PainPoint__c", "painPoint"), ("TargetSpec__c", "targetSpec"),
    ("DemoTargetDate__c", "demoTargetDate"), ("QualTargetDate__c", "qualTargetDate"),
    ("Competitor__c", "competitor"), ("LeadSource", "leadSource"),
    ("SensingEventId__c", "sensingEventId"), ("Description", "description"),
]


@app.get("/api/export-salesforce")
def export_salesforce(bu: str = "ALL", channel: str = "ALL", owner: str = "ALL"):
    from openpyxl import Workbook

    rows = [
        o for o in STORE["opportunities"]["opportunities"]
        if (bu == "ALL" or o.get("businessUnit") == bu)
        and (channel == "ALL" or o.get("channel") == channel)
        and (owner == "ALL" or o.get("owner") == owner)
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunity_Import"
    ws.append([c for c, _ in SF_COLS])
    for o in rows:
        ws.append([o.get(k) if o.get(k) is not None else "" for _, k in SF_COLS])
    ws2 = wb.create_sheet("참고")
    ws2.append(["항목", "내용"])
    ws2.append(["생성 기준", f"필터 사업부={bu}, 채널={channel}, 담당={owner} · {len(rows)}건"])
    ws2.append(["용도", "Salesforce Data Loader / Import Wizard용. 필드 매핑 규칙은 excel/salesforce_import_sample.xlsx '매핑표' 시트 참조"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"salesforce_import_{bu}_{channel}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# 프론트엔드 빌드가 있으면 루트에서 정적 서빙 (SPA fallback 포함)
if DIST_DIR.exists():
    app.mount("/", StaticFiles(directory=DIST_DIR, html=True), name="frontend")
